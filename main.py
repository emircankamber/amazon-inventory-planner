from fastapi import FastAPI, Form
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from typing import List, Tuple, Optional
from datetime import date
from io import BytesIO
import sqlite3
import math
import statistics
import json

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# -----------------------
# Config
# -----------------------
DB_NAME = "amazon_planner.db"

app = FastAPI()


# -----------------------
# DB helpers (SQLite)
# -----------------------
def get_conn():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS products (
        sku TEXT PRIMARY KEY,
        name TEXT,
        lead_time_days INTEGER NOT NULL,
        z_value REAL NOT NULL,
        fba_stock INTEGER NOT NULL DEFAULT 0,
        inbound_stock INTEGER NOT NULL DEFAULT 0,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP
    );
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS monthly_sales (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sku TEXT NOT NULL,
        year INTEGER NOT NULL,
        month INTEGER NOT NULL CHECK(month BETWEEN 1 AND 12),
        units_sold INTEGER NOT NULL,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        UNIQUE(sku, year, month),
        FOREIGN KEY(sku) REFERENCES products(sku)
    );
    """)

    conn.commit()
    conn.close()


@app.on_event("startup")
def startup():
    init_db()


# -----------------------
# Date helpers
# -----------------------
def last_n_calendar_months(n: int) -> List[Tuple[int, int]]:
    """Returns [(year, month), ...] for last n calendar months including current month."""
    today = date.today()
    y, m = today.year, today.month
    out = []
    for _ in range(n):
        out.append((y, m))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return out


def month_label(y: int, m: int) -> str:
    return f"{y}-{m:02d}"


# -----------------------
# Calculation engine
# -----------------------
def compute_from_last_months(
    lead_time: int,
    z: float,
    monthly_units: List[int],
    fba_stock: int,
    inbound_stock: int
) -> dict:
    """
    monthly_units: last 3 calendar months sales (found values only).
    """
    if not monthly_units:
        return {
            "daily_velocity": 0.0,
            "std_daily": 0.0,
            "safety_stock": 0.0,
            "rop": 0.0,
            "order_qty": 0.0,
        }

    mean_month = sum(monthly_units) / len(monthly_units)
    daily_velocity = mean_month / 30.0

    if len(monthly_units) < 2:
        std_daily = 0.0
    else:
        std_daily = statistics.stdev(monthly_units) / 30.0

    safety_stock = z * std_daily * math.sqrt(lead_time)
    rop = daily_velocity * lead_time + safety_stock
    order_qty = max(0.0, daily_velocity * 60 + safety_stock - (fba_stock + inbound_stock))

    return {
        "daily_velocity": daily_velocity,
        "std_daily": std_daily,
        "safety_stock": safety_stock,
        "rop": rop,
        "order_qty": order_qty,
    }


# -----------------------
# Excel helpers (openpyxl)
# -----------------------
def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(45, max(10, max_len + 2))


def workbook_to_stream(wb: Workbook) -> BytesIO:
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# -----------------------
# DB operations
# -----------------------
def upsert_product(cur, sku: str, name: str, lead_time_days: int, z_value: float, fba_stock: int, inbound_stock: int):
    cur.execute("""
    INSERT INTO products(sku, name, lead_time_days, z_value, fba_stock, inbound_stock)
    VALUES(?,?,?,?,?,?)
    ON CONFLICT(sku) DO UPDATE SET
      name=excluded.name,
      lead_time_days=excluded.lead_time_days,
      z_value=excluded.z_value,
      fba_stock=excluded.fba_stock,
      inbound_stock=excluded.inbound_stock,
      updated_at=CURRENT_TIMESTAMP;
    """, (sku, name, lead_time_days, z_value, fba_stock, inbound_stock))


def upsert_monthly_sales(cur, sku: str, years: List[int], months: List[int], units_sold: List[int]):
    for y, m, u in zip(years, months, units_sold):
        cur.execute("""
        INSERT INTO monthly_sales(sku, year, month, units_sold)
        VALUES(?,?,?,?)
        ON CONFLICT(sku, year, month) DO UPDATE SET
          units_sold=excluded.units_sold,
          created_at=CURRENT_TIMESTAMP;
        """, (sku, int(y), int(m), int(u)))


def fetch_product(cur, sku: str):
    return cur.execute("SELECT * FROM products WHERE sku=?", (sku,)).fetchone()


def fetch_month_units(cur, sku: str, y: int, m: int) -> Optional[int]:
    row = cur.execute("""
      SELECT units_sold FROM monthly_sales
      WHERE sku=? AND year=? AND month=?
    """, (sku, y, m)).fetchone()
    return None if row is None else int(row["units_sold"])


def compute_for_sku(cur, sku: str):
    prod = fetch_product(cur, sku)
    if prod is None:
        return None

    last3 = last_n_calendar_months(3)
    monthly_units_found: List[int] = []
    for y, m in last3:
        u = fetch_month_units(cur, sku, y, m)
        if u is not None:
            monthly_units_found.append(u)

    res = compute_from_last_months(
        lead_time=int(prod["lead_time_days"]),
        z=float(prod["z_value"]),
        monthly_units=monthly_units_found,
        fba_stock=int(prod["fba_stock"]),
        inbound_stock=int(prod["inbound_stock"])
    )

    return prod, last3, monthly_units_found, res


# -----------------------
# UI templating (NO f-string in JS-heavy blocks)
# -----------------------
def nav_html(active: str) -> str:
    def cls(name):
        base = "px-4 py-2 rounded-2xl text-sm border"
        if name == active:
            return base + " bg-indigo-500/20 border-indigo-400/30 text-indigo-100"
        return base + " bg-slate-900/40 border-slate-700/60 text-slate-200 hover:bg-slate-900/60"

    return f"""
    <div class="flex gap-2 flex-wrap">
      <a class="{cls('home')}" href="/">Giriş</a>
      <a class="{cls('products')}" href="/products">SKU Listesi</a>
      <a class="{cls('plan')}" href="/plan">Sipariş Listesi</a>
    </div>
    """


PAGE_SHELL = """
<!doctype html>
<html lang="tr">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <script src="https://cdn.tailwindcss.com"></script>
  <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
  <title>__TITLE__</title>
</head>
<body class="min-h-screen bg-gradient-to-br from-slate-950 via-slate-900 to-slate-950 text-slate-100">
  <div class="max-w-6xl mx-auto px-6 py-10">
    <div class="flex items-start justify-between gap-6 flex-wrap">
      <div>
        <h1 class="text-3xl md:text-4xl font-bold tracking-tight">__TITLE__</h1>
        <p class="text-slate-300 mt-2">Takvim ayı bazlı satış → son 3 ay ile planlama + son 6 ay trend.</p>
      </div>
      __NAV__
    </div>

    <div class="mt-8">
      __BODY__
    </div>
  </div>
</body>
</html>
"""


def page_shell(title: str, active_nav: str, body_html: str) -> str:
    html = PAGE_SHELL.replace("__TITLE__", title)
    html = html.replace("__NAV__", nav_html(active_nav))
    html = html.replace("__BODY__", body_html)
    return html


def build_default_rows_html() -> str:
    defaults = last_n_calendar_months(3)
    rows = []
    for y, m in defaults:
        rows.append(f"""
        <tr class="border-b border-slate-700/50">
          <td class="py-2">
            <input class="w-full rounded-xl bg-slate-900/60 border border-slate-700 px-3 py-2"
                   name="years" type="number" value="{y}" required>
          </td>
          <td class="py-2">
            <input class="w-full rounded-xl bg-slate-900/60 border border-slate-700 px-3 py-2"
                   name="months" type="number" min="1" max="12" value="{m}" required>
          </td>
          <td class="py-2">
            <input class="w-full rounded-xl bg-slate-900/60 border border-slate-700 px-3 py-2"
                   name="units_sold" type="number" min="0" placeholder="Adet" required>
          </td>
          <td class="py-2 text-right">
            <button type="button"
                    class="remove-row text-sm px-3 py-2 rounded-xl bg-rose-500/15 text-rose-200 border border-rose-500/30 hover:bg-rose-500/25">
              Sil
            </button>
          </td>
        </tr>
        """)
    return "\n".join(rows)


HOME_BODY = """
<form method="post" action="/upsert" class="grid grid-cols-1 lg:grid-cols-3 gap-6">
  <!-- Ürün kartı -->
  <div class="lg:col-span-1 rounded-3xl border border-slate-700/60 bg-slate-900/40 p-6 shadow-lg">
    <h2 class="text-xl font-semibold">Ürün Bilgileri</h2>

    <div class="mt-4 space-y-3">
      <label class="block">
        <span class="text-sm text-slate-300">SKU</span>
        <input name="sku" required
          class="mt-1 w-full rounded-2xl bg-slate-900/60 border border-slate-700 px-4 py-3 focus:outline-none focus:ring-2 focus:ring-indigo-500"
          placeholder="Örn: ABC-123"/>
      </label>

      <label class="block">
        <span class="text-sm text-slate-300">Ürün Adı</span>
        <input name="name"
          class="mt-1 w-full rounded-2xl bg-slate-900/60 border border-slate-700 px-4 py-3"
          placeholder="Opsiyonel"/>
      </label>

      <div class="grid grid-cols-2 gap-3">
        <label class="block">
          <span class="text-sm text-slate-300">Tedarik Süresi (gün)</span>
          <input name="lead_time_days" type="number" min="1" required
            class="mt-1 w-full rounded-2xl bg-slate-900/60 border border-slate-700 px-4 py-3"
            placeholder="Örn: 35"/>
        </label>

        <label class="block">
          <span class="text-sm text-slate-300">Z Katsayısı</span>
          <input name="z_value" type="number" step="0.01" required
            class="mt-1 w-full rounded-2xl bg-slate-900/60 border border-slate-700 px-4 py-3"
            placeholder="Örn: 1.65"/>
        </label>
      </div>

      <div class="grid grid-cols-2 gap-3">
        <label class="block">
          <span class="text-sm text-slate-300">FBA Stok</span>
          <input name="fba_stock" type="number" min="0" value="0" required
            class="mt-1 w-full rounded-2xl bg-slate-900/60 border border-slate-700 px-4 py-3"/>
        </label>

        <label class="block">
          <span class="text-sm text-slate-300">Yoldaki Stok</span>
          <input name="inbound_stock" type="number" min="0" value="0" required
            class="mt-1 w-full rounded-2xl bg-slate-900/60 border border-slate-700 px-4 py-3"/>
        </label>
      </div>
    </div>

    <button type="submit"
      class="mt-6 w-full rounded-2xl bg-indigo-500/90 hover:bg-indigo-500 px-5 py-3 font-semibold shadow-lg shadow-indigo-500/20">
      Kaydet + Hesapla
    </button>

    <p class="mt-3 text-xs text-slate-400">
      Not: Aynı SKU için aynı yıl/ay tekrar girersen o ay güncellenir.
    </p>
  </div>

  <!-- Aylık satışlar -->
  <div class="lg:col-span-2 rounded-3xl border border-slate-700/60 bg-slate-900/40 p-6 shadow-lg">
    <div class="flex items-center justify-between gap-3 flex-wrap">
      <h2 class="text-xl font-semibold">Aylık Satış (Çoklu Giriş)</h2>
      <button type="button" id="addRow"
        class="rounded-2xl bg-emerald-500/15 text-emerald-200 border border-emerald-500/30 hover:bg-emerald-500/25 px-4 py-2 text-sm font-semibold">
        + Ay Ekle
      </button>
    </div>

    <div class="mt-4 overflow-x-auto">
      <table class="w-full text-sm">
        <thead class="text-slate-300">
          <tr class="border-b border-slate-700/60">
            <th class="text-left py-2 pr-2">Yıl</th>
            <th class="text-left py-2 pr-2">Ay</th>
            <th class="text-left py-2 pr-2">Satış (adet)</th>
            <th class="text-right py-2">İşlem</th>
          </tr>
        </thead>
        <tbody id="salesBody">
          __ROWS__
        </tbody>
      </table>
    </div>

    <div class="mt-4 text-xs text-slate-400">
      Hesap: <b>son 3 takvim ayı</b>. Grafik: <b>son 6 ay</b>.
    </div>
  </div>
</form>

<script>
  const salesBody = document.getElementById("salesBody");
  const addBtn = document.getElementById("addRow");

  function attachRemoveHandlers() {
    document.querySelectorAll(".remove-row").forEach(btn => {
      btn.onclick = (e) => {
        const tr = e.target.closest("tr");
        if (salesBody.querySelectorAll("tr").length > 1) tr.remove();
      };
    });
  }

  addBtn.addEventListener("click", () => {
    const now = new Date();
    const y = now.getFullYear();
    const m = now.getMonth() + 1;

    const tr = document.createElement("tr");
    tr.className = "border-b border-slate-700/50";
    tr.innerHTML = `
      <td class="py-2">
        <input class="w-full rounded-xl bg-slate-900/60 border border-slate-700 px-3 py-2"
               name="years" type="number" value="${y}" required>
      </td>
      <td class="py-2">
        <input class="w-full rounded-xl bg-slate-900/60 border border-slate-700 px-3 py-2"
               name="months" type="number" min="1" max="12" value="${m}" required>
      </td>
      <td class="py-2">
        <input class="w-full rounded-xl bg-slate-900/60 border border-slate-700 px-3 py-2"
               name="units_sold" type="number" min="0" placeholder="Adet" required>
      </td>
      <td class="py-2 text-right">
        <button type="button"
                class="remove-row text-sm px-3 py-2 rounded-xl bg-rose-500/15 text-rose-200 border border-rose-500/30 hover:bg-rose-500/25">
          Sil
        </button>
      </td>`;
    salesBody.prepend(tr);
    attachRemoveHandlers();
  });

  attachRemoveHandlers();
</script>
"""


PRODUCT_DETAIL_TEMPLATE = """
<div class="grid grid-cols-1 lg:grid-cols-3 gap-6">
  <div class="lg:col-span-1 rounded-3xl border border-slate-700/60 bg-slate-900/40 p-6">
    <div class="text-slate-300 text-sm">SKU</div>
    <div class="text-2xl font-bold mt-1">__SKU__</div>
    <div class="text-slate-400 text-sm mt-1">__NAME__</div>

    <div class="mt-4 text-sm text-slate-300 space-y-1">
      <div>Lead Time: <b>__LT__</b> gün</div>
      <div>Z: <b>__Z__</b></div>
      <div>FBA: <b>__FBA__</b></div>
      <div>Yoldaki: <b>__INB__</b></div>
    </div>

    <div class="mt-5 rounded-2xl border border-slate-700/60 bg-slate-950/30 p-4">
      <div class="text-slate-300 text-sm">60 Gün Sipariş Önerisi</div>
      <div class="text-4xl font-extrabold mt-1">__ORDER__</div>
      <div class="text-xs text-slate-400 mt-1">max(0, H*60 + SS - (FBA+Yoldaki))</div>
    </div>

    <div class="mt-4 grid grid-cols-2 gap-3">
      <a href="/products"
         class="text-center rounded-2xl border border-slate-700/60 bg-slate-900/40 hover:bg-slate-900/60 px-4 py-2 text-sm">SKU Listesi</a>
      <a href="/plan"
         class="text-center rounded-2xl border border-slate-700/60 bg-slate-900/40 hover:bg-slate-900/60 px-4 py-2 text-sm">Sipariş Listesi</a>
    </div>
  </div>

  <div class="lg:col-span-2 space-y-6">
    <div class="rounded-3xl border border-slate-700/60 bg-slate-900/40 p-6">
      <div class="text-slate-300 text-sm">Son 3 Takvim Ayı (Hesap Bazı)</div>
      <div class="text-slate-200 text-sm mt-1">
        Aylar: <b>__LAST3_LABELS__</b> |
        Bulunan satışlar: <b>__LAST3_FOUND__</b>
      </div>

      <div class="mt-5 grid grid-cols-1 md:grid-cols-4 gap-4">
        <div class="rounded-2xl border border-slate-700/60 bg-slate-950/30 p-4">
          <div class="text-slate-300 text-xs">Günlük Hız</div>
          <div class="text-2xl font-bold mt-1">__H__</div>
        </div>
        <div class="rounded-2xl border border-slate-700/60 bg-slate-950/30 p-4">
          <div class="text-slate-300 text-xs">Std. Sapma</div>
          <div class="text-2xl font-bold mt-1">__STD__</div>
        </div>
        <div class="rounded-2xl border border-slate-700/60 bg-slate-950/30 p-4">
          <div class="text-slate-300 text-xs">Güvenlik Stoku</div>
          <div class="text-2xl font-bold mt-1">__SS__</div>
        </div>
        <div class="rounded-2xl border border-slate-700/60 bg-slate-950/30 p-4">
          <div class="text-slate-300 text-xs">ROP</div>
          <div class="text-2xl font-bold mt-1">__ROP__</div>
        </div>
      </div>
    </div>

    <div class="rounded-3xl border border-slate-700/60 bg-slate-900/40 p-6">
      <div class="text-slate-300 text-sm">Son 6 Ay Satış Trend</div>
      <div class="mt-4">
        <canvas id="salesChart" height="120"></canvas>
      </div>
      <div class="text-xs text-slate-400 mt-3">Eksik aylar grafikte 0 gösterilir.</div>
    </div>
  </div>
</div>

<script>
  const labels = __LABELS_JSON__;
  const data = __DATA_JSON__;

  const ctx = document.getElementById('salesChart').getContext('2d');
  new Chart(ctx, {
    type: 'line',
    data: {
      labels,
      datasets: [{
        label: 'Aylık Satış (adet)',
        data,
        tension: 0.3
      }]
    },
    options: {
      responsive: true,
      plugins: {
        legend: {
          labels: { color: '#cbd5e1' }
        }
      },
      scales: {
        x: { ticks: { color: '#94a3b8' }, grid: { color: 'rgba(148,163,184,0.15)' } },
        y: { ticks: { color: '#94a3b8' }, grid: { color: 'rgba(148,163,184,0.15)' } }
      }
    }
  });
</script>
"""


# -----------------------
# Pages
# -----------------------
@app.get("/", response_class=HTMLResponse)
def home():
    body = HOME_BODY.replace("__ROWS__", build_default_rows_html())
    return page_shell("Amazon Stok Planlama", "home", body)


@app.post("/upsert", response_class=HTMLResponse)
def upsert(
    sku: str = Form(...),
    name: str = Form(""),
    lead_time_days: int = Form(...),
    z_value: float = Form(...),
    fba_stock: int = Form(0),
    inbound_stock: int = Form(0),
    years: List[int] = Form(...),
    months: List[int] = Form(...),
    units_sold: List[int] = Form(...)
):
    if not (len(years) == len(months) == len(units_sold)):
        return HTMLResponse("<h2>Hata</h2><p>Aylık satış satırlarında eksik alan var.</p><a href='/'>Geri</a>", status_code=400)

    conn = get_conn()
    cur = conn.cursor()

    upsert_product(cur, sku, name, lead_time_days, z_value, fba_stock, inbound_stock)
    upsert_monthly_sales(cur, sku, years, months, units_sold)

    conn.commit()
    conn.close()

    return RedirectResponse(url=f"/product/{sku}", status_code=303)


@app.get("/product/{sku}", response_class=HTMLResponse)
def product_detail(sku: str):
    conn = get_conn()
    cur = conn.cursor()

    computed = compute_for_sku(cur, sku)
    if computed is None:
        conn.close()
        return page_shell("Ürün Bulunamadı", "products",
                          f"<div class='rounded-3xl border border-slate-700/60 bg-slate-900/40 p-6'>SKU bulunamadı: <b>{sku}</b></div>")

    prod, last3, monthly_found, res = computed

    # last 6 months (old -> new), missing months as 0 for continuity
    last6 = list(reversed(last_n_calendar_months(6)))
    labels6 = [month_label(y, m) for y, m in last6]
    units6 = [(fetch_month_units(cur, sku, y, m) or 0) for y, m in last6]

    conn.close()

    template = PRODUCT_DETAIL_TEMPLATE
    template = template.replace("__SKU__", sku)
    template = template.replace("__NAME__", (prod["name"] or ""))
    template = template.replace("__LT__", str(int(prod["lead_time_days"])))
    template = template.replace("__Z__", str(float(prod["z_value"])))
    template = template.replace("__FBA__", str(int(prod["fba_stock"])))
    template = template.replace("__INB__", str(int(prod["inbound_stock"])))
    template = template.replace("__ORDER__", str(int(round(res["order_qty"]))))

    template = template.replace("__LAST3_LABELS__", ", ".join(month_label(y, m) for y, m in last3))
    template = template.replace("__LAST3_FOUND__", str(monthly_found if monthly_found else "yok"))

    template = template.replace("__H__", f"{res['daily_velocity']:.2f}")
    template = template.replace("__STD__", f"{res['std_daily']:.2f}")
    template = template.replace("__SS__", f"{res['safety_stock']:.2f}")
    template = template.replace("__ROP__", f"{res['rop']:.2f}")

    template = template.replace("__LABELS_JSON__", json.dumps(labels6))
    template = template.replace("__DATA_JSON__", json.dumps(units6))

    return page_shell(f"Ürün Detay — {sku}", "products", template)


@app.get("/products", response_class=HTMLResponse)
def products():
    conn = get_conn()
    cur = conn.cursor()
    skus = [r["sku"] for r in cur.execute("SELECT sku FROM products ORDER BY updated_at DESC").fetchall()]

    rows = []
    for sku in skus:
        prod, _, _, res = compute_for_sku(cur, sku)
        order_qty = int(round(res["order_qty"]))
        badge = "bg-emerald-500/15 text-emerald-200 border-emerald-500/30"
        if order_qty > 0:
            badge = "bg-rose-500/15 text-rose-200 border-rose-500/30"

        rows.append(f"""
        <tr class="border-b border-slate-700/50">
          <td class="py-3 pr-3 font-semibold">{sku}</td>
          <td class="py-3 pr-3 text-slate-300">{prod["name"] or ""}</td>
          <td class="py-3 pr-3 text-right">{res["daily_velocity"]:.2f}</td>
          <td class="py-3 pr-3 text-right">{res["rop"]:.2f}</td>
          <td class="py-3 pr-3 text-right">
            <span class="px-3 py-1 rounded-2xl border {badge}">{order_qty}</span>
          </td>
          <td class="py-3 text-right">
            <a class="px-4 py-2 rounded-2xl text-sm border border-slate-700/60 bg-slate-900/40 hover:bg-slate-900/60"
               href="/product/{sku}">Detay</a>
          </td>
        </tr>
        """)

    conn.close()

    body = f"""
    <div class="rounded-3xl border border-slate-700/60 bg-slate-900/40 p-6 shadow-lg">
      <div class="flex items-end justify-between flex-wrap gap-3">
        <div>
          <h2 class="text-xl font-semibold">SKU Listesi</h2>
          <p class="text-slate-400 text-sm mt-1">Sipariş > 0 olanlar kırmızı görünür.</p>
        </div>

        <div class="flex gap-2 flex-wrap">
          <a href="/"
             class="rounded-2xl bg-indigo-500/90 hover:bg-indigo-500 px-4 py-2 text-sm font-semibold">+ Yeni Giriş</a>
          <a href="/export/products.xlsx"
             class="rounded-2xl border border-slate-700/60 bg-slate-900/40 hover:bg-slate-900/60 px-4 py-2 text-sm font-semibold">
             Excel indir (SKU)
          </a>
          <a href="/export/plan.xlsx"
             class="rounded-2xl border border-slate-700/60 bg-slate-900/40 hover:bg-slate-900/60 px-4 py-2 text-sm font-semibold">
             Excel indir (Sipariş)
          </a>
        </div>
      </div>

      <div class="mt-5 overflow-x-auto">
        <table class="w-full text-sm">
          <thead class="text-slate-300">
            <tr class="border-b border-slate-700/60">
              <th class="text-left py-2 pr-3">SKU</th>
              <th class="text-left py-2 pr-3">Ürün</th>
              <th class="text-right py-2 pr-3">Günlük Hız</th>
              <th class="text-right py-2 pr-3">ROP</th>
              <th class="text-right py-2 pr-3">Sipariş (60g)</th>
              <th class="text-right py-2"> </th>
            </tr>
          </thead>
          <tbody>
            {''.join(rows) if rows else '<tr><td class="py-4 text-slate-400" colspan="6">Henüz ürün yok.</td></tr>'}
          </tbody>
        </table>
      </div>
    </div>
    """
    return page_shell("SKU Listesi", "products", body)


@app.get("/plan", response_class=HTMLResponse)
def plan():
    conn = get_conn()
    cur = conn.cursor()
    skus = [r["sku"] for r in cur.execute("SELECT sku FROM products ORDER BY updated_at DESC").fetchall()]

    rows = []
    for sku in skus:
        prod, _, _, res = compute_for_sku(cur, sku)
        order_qty = int(round(res["order_qty"]))
        if order_qty <= 0:
            continue

        rows.append(f"""
        <tr class="border-b border-slate-700/50">
          <td class="py-3 pr-3 font-semibold">{sku}</td>
          <td class="py-3 pr-3 text-slate-300">{prod["name"] or ""}</td>
          <td class="py-3 pr-3 text-right">{int(prod["fba_stock"])}</td>
          <td class="py-3 pr-3 text-right">{int(prod["inbound_stock"])}</td>
          <td class="py-3 pr-3 text-right">{res["rop"]:.2f}</td>
          <td class="py-3 pr-3 text-right">
            <span class="px-3 py-1 rounded-2xl border bg-rose-500/15 text-rose-200 border-rose-500/30">{order_qty}</span>
          </td>
          <td class="py-3 text-right">
            <a class="px-4 py-2 rounded-2xl text-sm border border-slate-700/60 bg-slate-900/40 hover:bg-slate-900/60"
               href="/product/{sku}">Detay</a>
          </td>
        </tr>
        """)

    conn.close()

    body = f"""
    <div class="rounded-3xl border border-slate-700/60 bg-slate-900/40 p-6 shadow-lg">
      <div class="flex items-end justify-between flex-wrap gap-3">
        <div>
          <h2 class="text-xl font-semibold">Sipariş Listesi (Order > 0)</h2>
          <p class="text-slate-400 text-sm mt-1">2 aylık plan: H*60 + SS - (FBA+Yoldaki)</p>
        </div>

        <div class="flex gap-2 flex-wrap">
          <a href="/"
             class="rounded-2xl bg-indigo-500/90 hover:bg-indigo-500 px-4 py-2 text-sm font-semibold">+ Yeni Giriş</a>
          <a href="/export/plan.xlsx"
             class="rounded-2xl border border-slate-700/60 bg-slate-900/40 hover:bg-slate-900/60 px-4 py-2 text-sm font-semibold">
             Excel indir (Sipariş)
          </a>
        </div>
      </div>

      <div class="mt-5 overflow-x-auto">
        <table class="w-full text-sm">
          <thead class="text-slate-300">
            <tr class="border-b border-slate-700/60">
              <th class="text-left py-2 pr-3">SKU</th>
              <th class="text-left py-2 pr-3">Ürün</th>
              <th class="text-right py-2 pr-3">FBA</th>
              <th class="text-right py-2 pr-3">Yoldaki</th>
              <th class="text-right py-2 pr-3">ROP</th>
              <th class="text-right py-2 pr-3">Sipariş</th>
              <th class="text-right py-2"> </th>
            </tr>
          </thead>
          <tbody>
            {''.join(rows) if rows else '<tr><td class="py-4 text-slate-400" colspan="7">Sipariş gereken ürün yok.</td></tr>'}
          </tbody>
        </table>
      </div>
    </div>
    """
    return page_shell("Sipariş Listesi", "plan", body)


# -----------------------
# Excel export endpoints
# -----------------------
@app.get("/export/products.xlsx")
def export_products_xlsx():
    conn = get_conn()
    cur = conn.cursor()
    skus = [r["sku"] for r in cur.execute("SELECT sku FROM products ORDER BY updated_at DESC").fetchall()]

    wb = Workbook()
    ws = wb.active
    ws.title = "SKUListesi"

    headers = ["SKU", "Ürün", "LeadTime", "Z", "FBA", "Yoldaki", "GünlükHız", "Std", "SS", "ROP", "Sipariş(60g)"]
    ws.append(headers)

    for sku in skus:
        prod, _, _, res = compute_for_sku(cur, sku)
        ws.append([
            sku,
            prod["name"] or "",
            int(prod["lead_time_days"]),
            float(prod["z_value"]),
            int(prod["fba_stock"]),
            int(prod["inbound_stock"]),
            round(res["daily_velocity"], 4),
            round(res["std_daily"], 4),
            round(res["safety_stock"], 4),
            round(res["rop"], 4),
            int(round(res["order_qty"]))
        ])

    conn.close()
    autosize_columns(ws)

    bio = workbook_to_stream(wb)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="sku_listesi.xlsx"'}
    )


@app.get("/export/plan.xlsx")
def export_plan_xlsx():
    conn = get_conn()
    cur = conn.cursor()
    skus = [r["sku"] for r in cur.execute("SELECT sku FROM products ORDER BY updated_at DESC").fetchall()]

    wb = Workbook()
    ws = wb.active
    ws.title = "SiparisListesi"

    headers = ["SKU", "Ürün", "LeadTime", "Z", "FBA", "Yoldaki", "GünlükHız", "ROP", "Sipariş(60g)"]
    ws.append(headers)

    for sku in skus:
        prod, _, _, res = compute_for_sku(cur, sku)
        order_qty = int(round(res["order_qty"]))
        if order_qty <= 0:
            continue
        ws.append([
            sku,
            prod["name"] or "",
            int(prod["lead_time_days"]),
            float(prod["z_value"]),
            int(prod["fba_stock"]),
            int(prod["inbound_stock"]),
            round(res["daily_velocity"], 4),
            round(res["rop"], 4),
            order_qty
        ])

    conn.close()
    autosize_columns(ws)

    bio = workbook_to_stream(wb)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="siparis_listesi.xlsx"'}
    )


# -----------------------
# Google Sheets push (READY, needs credentials.json)
# -----------------------
"""
Google Sheets'e otomatik yazdırma için 1 kez kurulum gerekir:

1) Google Cloud Console:
   - Sheets API enable
   - Service Account oluştur
   - credentials.json indir

2) Bu proje klasörüne credentials.json koy.

3) Yeni Sheet oluşturup service account email'ine edit yetkisi ver
   (veya endpoint yeni sheet oluşturacak şekilde de genişletilebilir)

Aşağıdaki endpoint'i aktif etmek için:
pip install google-api-python-client google-auth
"""

# from google.oauth2.service_account import Credentials
# from googleapiclient.discovery import build

# GOOGLE_CREDS_FILE = "credentials.json"
# GOOGLE_SHEET_ID = "BURAYA_SHEET_ID"  # örn: 1AbC... (URL içinden)

# def get_sheets_service():
#     scopes = ["https://www.googleapis.com/auth/spreadsheets"]
#     creds = Credentials.from_service_account_file(GOOGLE_CREDS_FILE, scopes=scopes)
#     return build("sheets", "v4", credentials=creds)

# @app.post("/push/plan/google-sheet")
# def push_plan_to_google_sheet():
#     """
#     Sipariş listesini Google Sheets'e yazar.
#     Not: GOOGLE_SHEET_ID ve credentials.json hazır olmalı.
#     """
#     service = get_sheets_service()
#
#     conn = get_conn()
#     cur = conn.cursor()
#     skus = [r["sku"] for r in cur.execute("SELECT sku FROM products ORDER BY updated_at DESC").fetchall()]
#
#     values = [["SKU", "Ürün", "LeadTime", "Z", "FBA", "Yoldaki", "GünlükHız", "ROP", "Sipariş(60g)"]]
#     for sku in skus:
#         prod, _, _, res = compute_for_sku(cur, sku)
#         order_qty = int(round(res["order_qty"]))
#         if order_qty <= 0:
#             continue
#         values.append([
#             sku,
#             prod["name"] or "",
#             int(prod["lead_time_days"]),
#             float(prod["z_value"]),
#             int(prod["fba_stock"]),
#             int(prod["inbound_stock"]),
#             round(res["daily_velocity"], 4),
#             round(res["rop"], 4),
#             order_qty
#         ])
#
#     conn.close()
#
#     # write to sheet (range example: "Plan!A1")
#     body = {"values": values}
#     service.spreadsheets().values().update(
#         spreadsheetId=GOOGLE_SHEET_ID,
#         range="Plan!A1",
#         valueInputOption="RAW",
#         body=body
#     ).execute()
#
#     return {"status": "ok", "rows_written": len(values), "sheet_id": GOOGLE_SHEET_ID}
